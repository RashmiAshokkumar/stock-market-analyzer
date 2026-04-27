# src/report_generator.py

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import os
import sys

# Add parent directory to path so we can import our own modules
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from technical_analysis import full_technical_analysis, load_prices
from sqlalchemy import create_engine, text

BASE_DIR_DB = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DB_PATH_RPT = os.path.join(BASE_DIR_DB, "database", "stocks.db")
engine_rpt  = create_engine(f"sqlite:///{DB_PATH_RPT}")

def load_full_prices(ticker: str) -> pd.DataFrame:
    """Loads all OHLCV columns for a ticker."""
    with engine_rpt.connect() as conn:
        df = pd.read_sql(
            text("SELECT date, open, high, low, close, volume FROM stock_prices WHERE ticker = :t ORDER BY date"),
            conn, params={"t": ticker}
        )
    df["date"] = pd.to_datetime(df["date"]).dt.date
    return df

from anomaly_detection   import detect_anomalies
from screener            import build_fundamentals_df, week52_position

# ── Output path ────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
OUTPUT_PATH = os.path.join(BASE_DIR, "reports", "weekly_report.xlsx")

# ── Colour palette ─────────────────────────────────────────────────────────────
# openpyxl uses hex colours without the # symbol
DARK_GREEN  = "1D6A3A"
MID_GREEN   = "C6EFCE"
DARK_RED    = "9C0006"
MID_RED     = "FFC7CE"
DARK_YELLOW = "7D6608"
MID_YELLOW  = "FFEB9C"
HEADER_BG   = "1F4E79"   # dark blue for headers
HEADER_FG   = "FFFFFF"   # white text on headers
ALT_ROW     = "F2F7FF"   # light blue for alternating rows
BORDER_COL  = "B8CCE4"

TICKERS = ["AAPL", "MSFT", "GOOGL", "TSLA", "JPM"]


# ── Style helpers ──────────────────────────────────────────────────────────────
def header_font():
    return Font(name="Calibri", bold=True, color=HEADER_FG, size=11)

def body_font():
    return Font(name="Calibri", size=10)

def bold_font():
    return Font(name="Calibri", bold=True, size=10)

def header_fill():
    return PatternFill("solid", fgColor=HEADER_BG)

def green_fill():
    return PatternFill("solid", fgColor=MID_GREEN)

def red_fill():
    return PatternFill("solid", fgColor=MID_RED)

def yellow_fill():
    return PatternFill("solid", fgColor=MID_YELLOW)

def alt_fill():
    return PatternFill("solid", fgColor=ALT_ROW)

def thin_border():
    side = Side(style="thin", color=BORDER_COL)
    return Border(left=side, right=side, top=side, bottom=side)

def write_headers(ws, headers: list, row: int = 1):
    """Writes a styled header row."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = header_font()
        cell.fill      = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()

def auto_width(ws, min_width=10, max_width=30):
    """Auto-sizes all columns based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, min_width), max_width
        )

def colour_cell(cell, value, green_above=None, red_below=None):
    """
    Colours a cell green if value >= green_above,
    red if value <= red_below, yellow otherwise.
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return
    if green_above is not None and value >= green_above:
        cell.fill = green_fill()
        cell.font = Font(name="Calibri", color=DARK_GREEN, size=10)
    elif red_below is not None and value <= red_below:
        cell.fill = red_fill()
        cell.font = Font(name="Calibri", color=DARK_RED, size=10)
    else:
        cell.fill = yellow_fill()
        cell.font = Font(name="Calibri", color=DARK_YELLOW, size=10)


# ── Sheet 1: Executive Summary ─────────────────────────────────────────────────
def write_summary_sheet(wb: Workbook, fundamentals_df: pd.DataFrame):
    """
    First sheet — a quick at-a-glance view of all stocks.
    Shows current price, 52-week position, P/E, and signals.
    """
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws["A1"] = "Stock Market Analyzer — Weekly Report"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color=HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells("A1:H1")

    ws["A2"] = f"Generated automatically by Python · {pd.Timestamp.now().strftime('%B %d, %Y')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="808080")
    ws.merge_cells("A2:H2")

    # Headers
    headers = [
        "Ticker", "Company", "Price", "52W Position %",
        "P/E Ratio", "ROE %", "Rev Growth %", "Market Cap ($B)"
    ]
    write_headers(ws, headers, row=4)

    # Data rows
    cols = [
        "ticker", "company", "current_price", "52w_position_%",
        "pe_ratio", "return_on_equity", "revenue_growth", "market_cap_B"
    ]
    for i, (_, row) in enumerate(fundamentals_df.iterrows(), start=5):
        fill = alt_fill() if i % 2 == 0 else PatternFill()
        for j, col in enumerate(cols, start=1):
            val  = row.get(col)
            cell = ws.cell(row=i, column=j, value=val)
            cell.font      = body_font()
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if fill.patternType:
                cell.fill = fill

            # Colour code 52W position
            if col == "52w_position_%":
                colour_cell(cell, val, green_above=60, red_below=30)

            # Colour code ROE
            if col == "return_on_equity":
                colour_cell(cell, val, green_above=20, red_below=10)

            # Colour code revenue growth
            if col == "revenue_growth":
                colour_cell(cell, val, green_above=10, red_below=0)

    auto_width(ws)


# ── Sheet 2: Price History ─────────────────────────────────────────────────────
def write_price_sheet(wb: Workbook):
    """
    Writes the last 60 days of OHLCV data for all tickers.
    Colours the daily return column green/red.
    """
    ws = wb.create_sheet("Price History")
    ws.sheet_view.showGridLines = False

    headers = ["Date", "Ticker", "Open", "High", "Low", "Close", "Volume", "Daily Return %"]
    write_headers(ws, headers)

    row_num = 2
    for ticker in TICKERS:
        df = load_full_prices(ticker).tail(60)
        df["daily_return"] = df["close"].pct_change() * 100

        for _, row in df.iterrows():
            values = [
                str(row["date"]), ticker,
                round(row["open"],  2), round(row["high"], 2),
                round(row["low"],   2), round(row["close"],2),
                int(row["volume"]),
                round(row["daily_return"], 2) if not pd.isna(row["daily_return"]) else None
            ]
            fill = alt_fill() if row_num % 2 == 0 else PatternFill()
            for col_num, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=val)
                cell.font      = body_font()
                cell.border    = thin_border()
                cell.alignment = Alignment(horizontal="center")
                if fill.patternType:
                    cell.fill = fill

            # Colour the return cell
            ret_cell = ws.cell(row=row_num, column=8)
            ret_val  = values[7]
            if ret_val is not None:
                colour_cell(ret_cell, ret_val, green_above=1.0, red_below=-1.0)

            row_num += 1

    auto_width(ws)


# ── Sheet 3: Technical Indicators ─────────────────────────────────────────────
def write_indicators_sheet(wb: Workbook):
    """
    Writes RSI, MACD, and Bollinger Band values for all tickers.
    Colours RSI cells based on overbought/oversold zones.
    """
    ws = wb.create_sheet("Technical Indicators")
    ws.sheet_view.showGridLines = False

    headers = [
        "Date", "Ticker", "Close", "RSI", "RSI Signal",
        "MACD", "MACD Signal", "Histogram", "MACD Trend",
        "BB Upper", "BB Lower", "%B", "BB Signal"
    ]
    write_headers(ws, headers)

    row_num = 2
    for ticker in TICKERS:
        df = full_technical_analysis(ticker).tail(30)

        for _, row in df.iterrows():
            values = [
                str(row["date"]), ticker, row.get("close"),
                row.get("rsi"),   str(row.get("rsi_signal",  "")),
                row.get("macd"),  row.get("signal"), row.get("histogram"),
                str(row.get("macd_signal", "")),
                row.get("bb_upper"), row.get("bb_lower"),
                row.get("bb_percent_b"), str(row.get("bb_signal", ""))
            ]
            fill = alt_fill() if row_num % 2 == 0 else PatternFill()
            for col_num, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=val)
                cell.font      = body_font()
                cell.border    = thin_border()
                cell.alignment = Alignment(horizontal="center")
                if fill.patternType:
                    cell.fill = fill

            # Colour RSI cell
            rsi_val  = row.get("rsi")
            rsi_cell = ws.cell(row=row_num, column=4)
            if rsi_val is not None:
                colour_cell(rsi_cell, rsi_val, green_above=70, red_below=30)

            # Colour MACD trend cell
            trend_cell = ws.cell(row=row_num, column=9)
            if str(row.get("macd_signal", "")) == "bullish":
                trend_cell.fill = green_fill()
                trend_cell.font = Font(name="Calibri", color=DARK_GREEN, size=10)
            else:
                trend_cell.fill = red_fill()
                trend_cell.font = Font(name="Calibri", color=DARK_RED, size=10)

            row_num += 1

    auto_width(ws)


# ── Sheet 4: Anomalies ─────────────────────────────────────────────────────────
def write_anomalies_sheet(wb: Workbook):
    """
    Lists all detected anomalies across all tickers.
    Colours by severity.
    """
    ws = wb.create_sheet("Anomalies")
    ws.sheet_view.showGridLines = False

    headers = ["Date", "Ticker", "Close", "Daily Return %", "Volume", "Z-Score", "Severity"]
    write_headers(ws, headers)

    row_num = 2
    for ticker in TICKERS:
        anomalies = detect_anomalies(ticker)
        for _, row in anomalies.iterrows():
            severity = str(row.get("severity", ""))
            values   = [
                str(row["date"]), ticker,
                row["close"], row["daily_return"],
                int(row["volume"]), row["zscore_daily_return"],
                severity
            ]
            for col_num, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=val)
                cell.font      = body_font()
                cell.border    = thin_border()
                cell.alignment = Alignment(horizontal="center")

            # Colour entire row by severity
            if severity == "extreme":
                row_fill = red_fill()
            elif severity == "moderate":
                row_fill = yellow_fill()
            else:
                row_fill = alt_fill() if row_num % 2 == 0 else PatternFill()

            if row_fill.patternType:
                for col_num in range(1, 8):
                    ws.cell(row=row_num, column=col_num).fill = row_fill

            row_num += 1

    auto_width(ws)


# ── Sheet 5: Fundamentals ──────────────────────────────────────────────────────
def write_fundamentals_sheet(wb: Workbook, fundamentals_df: pd.DataFrame):
    """
    Full fundamentals table with all ratios.
    """
    ws = wb.create_sheet("Fundamentals")
    ws.sheet_view.showGridLines = False

    headers = [
        "Ticker", "Company", "Sector", "Market Cap ($B)",
        "P/E Ratio", "Forward P/E", "Price/Book",
        "Debt/Equity", "ROE %", "Revenue Growth %",
        "Gross Margin %", "Current Price", "52W High", "52W Low"
    ]
    write_headers(ws, headers)

    cols = [
        "ticker", "company", "sector", "market_cap_B",
        "pe_ratio", "forward_pe", "price_to_book",
        "debt_to_equity", "return_on_equity", "revenue_growth",
        "gross_margins", "current_price", "52w_high", "52w_low"
    ]

    for i, (_, row) in enumerate(fundamentals_df.iterrows(), start=2):
        fill = alt_fill() if i % 2 == 0 else PatternFill()
        for j, col in enumerate(cols, start=1):
            val  = row.get(col)
            cell = ws.cell(row=i, column=j, value=val)
            cell.font      = body_font()
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if fill.patternType:
                cell.fill = fill

        # Colour P/E ratio
        pe_cell = ws.cell(row=i, column=5)
        colour_cell(pe_cell, row.get("pe_ratio"), red_below=15, green_above=25)

        # Colour ROE
        roe_cell = ws.cell(row=i, column=9)
        colour_cell(roe_cell, row.get("return_on_equity"), green_above=20, red_below=10)

    auto_width(ws)


# ── Sheet 6: Price chart ───────────────────────────────────────────────────────
def write_chart_sheet(wb: Workbook):
    """
    Creates a line chart of closing prices for all tickers.
    """
    ws_data = wb.create_sheet("ChartData")

    ws_data["A1"] = "Date"
    col = 2
    all_dates = None

    for ticker in TICKERS:
        df = load_full_prices(ticker).tail(60).reset_index(drop=True)
        if all_dates is None:
            for i, d in enumerate(df["date"], start=2):
                ws_data.cell(row=i, column=1, value=str(d))
            all_dates = df["date"]

        ws_data.cell(row=1, column=col, value=ticker)
        for i, price in enumerate(df["close"], start=2):
            ws_data.cell(row=i, column=col, value=round(price, 2))
        col += 1

    ws_chart = wb.create_sheet("Price Chart")
    ws_chart.sheet_view.showGridLines = False

    chart = LineChart()
    chart.title           = "60-Day Closing Prices"
    chart.style           = 10
    chart.y_axis.title    = "Price (USD)"
    chart.x_axis.title    = "Date"
    chart.width           = 35    # wider
    chart.height          = 20    # taller
    chart.y_axis.numFmt   = '#,##0.00'
    chart.y_axis.tickLblPos = "low"

    for i, ticker in enumerate(TICKERS, start=2):
        data = Reference(ws_data, min_col=i, min_row=1, max_row=61)
        chart.add_data(data, titles_from_data=True)

    # Add date labels on X axis
    dates = Reference(ws_data, min_col=1, min_row=2, max_row=61)
    chart.set_categories(dates)

    ws_chart.add_chart(chart, "A1")
    

# ── Main runner ────────────────────────────────────────────────────────────────
def generate_report(tickers: list = TICKERS):
    """
    Master function — generates the complete Excel report.
    Call this from the command line or from another script.
    """
    print("Generating weekly report...")

    print("  Fetching fundamentals...")
    fundamentals = build_fundamentals_df(tickers)
    fundamentals = week52_position(fundamentals)

    wb = Workbook()

    print("  Writing Summary sheet...")
    write_summary_sheet(wb, fundamentals)

    print("  Writing Price History sheet...")
    write_price_sheet(wb)

    print("  Writing Technical Indicators sheet...")
    write_indicators_sheet(wb)

    print("  Writing Anomalies sheet...")
    write_anomalies_sheet(wb)

    print("  Writing Fundamentals sheet...")
    write_fundamentals_sheet(wb, fundamentals)

    print("  Building Price Chart...")
    write_chart_sheet(wb)

    # Save the workbook
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nReport saved to: {OUTPUT_PATH}")
    print("Open it in Excel to view the full report.")


if __name__ == "__main__":
    generate_report()

    